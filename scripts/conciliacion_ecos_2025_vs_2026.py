"""
conciliacion_ecos_2025_vs_2026.py
─────────────────────────────────
Genera un reporte en Excel estético comparando ECOs activos en 2025 vs 2026.
Contiene:
  1. RESUMEN GENERAL: Conteo de unidades por mes en ambos años y la diferencia.
  2. COMPARATIVA GENERAL: Comparativa detallada para un mes seleccionado (default: Enero).
  3. PASE: Comparativa detallada de unidades activas en Pase para el mes seleccionado.
  4. SUPRAMAX: Comparativa detallada de unidades activas en Supramax para el mes seleccionado.
  5. EDENRED: Comparativa detallada de unidades activas en Edenred para el mes seleccionado.
  6. MANTENIMIENTOS: Comparativa detallada de unidades con mantenimientos (Google Sheets) para el mes seleccionado.

Uso:
    .venv/bin/python scripts/conciliacion_ecos_2025_vs_2026.py --mes 1
"""

import os
import sys
import argparse
import re
import pandas as pd
from google.cloud import bigquery
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

load_dotenv()

def _normalize_eco(val):
    s = str(val).strip().upper().replace(' ', '').replace('.', '')
    m = re.match(r'^(AU|CA)-?(\d{1,3})(?!\d)', s)
    if m:
        return f"{m.group(1)}-{m.group(2).zfill(3)}"
    return s

def get_original_eco_map():
    mapping = {}
    edenred_path = os.path.join(PROJECT_ROOT, "CONSOLIDADO_CRUDO_EDENRED.csv")
    if os.path.exists(edenred_path):
        try:
            df = pd.read_csv(edenred_path, usecols=lambda c: 'veh' in c.lower() or c == 'Vehículo' or 'vehã' in c.lower(), encoding='latin1', low_memory=False)
            col_name = df.columns[0]
            for val in df[col_name].dropna().unique():
                val_clean = str(val).strip()
                if not val_clean:
                    continue
                norm = _normalize_eco(val_clean)
                if norm != val_clean and 'LZC' in val_clean:
                    mapping[norm] = val_clean
        except Exception as e:
            print(f"⚠️ Error al leer mapeo original: {e}")
    return mapping

# --- CONFIGURACIÓN DE ESTILOS ESTÉTICOS ---
FONT_NAME = 'Century Gothic'
FONT_NORMAL = Font(name=FONT_NAME, size=11)
FONT_BOLD = Font(name=FONT_NAME, size=11, bold=True)
FONT_TITLE = Font(name=FONT_NAME, size=15, bold=True, color="1F4E78")
FONT_HEADER = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")

FILL_HEADER = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
FILL_ZEBRA = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_GREEN = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Soft green
FILL_RED = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")    # Soft red

BORDER_THIN = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')

def get_month_name(m):
    meses = {
        1: 'Enero', 2: 'Febrero', 3: 'Marzo', 4: 'Abril', 5: 'Mayo', 6: 'Junio',
        7: 'Julio', 8: 'Agosto', 9: 'Septiembre', 10: 'Octubre', 11: 'Noviembre', 12: 'Diciembre'
    }
    return meses.get(m, str(m))

def format_cell(cell, font=FONT_NORMAL, fill=None, alignment=ALIGN_LEFT, border=BORDER_THIN, num_format=None):
    cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if num_format:
        cell.number_format = num_format

def autofit_columns(ws, start_row=1):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        # Skip the title row for width calculation
        for cell in col[start_row-1:]:
            val = str(cell.value or '')
            if cell.number_format and ('$' in cell.number_format):
                val = f"${val}"
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def generate_comparative_df(df_mes, filter_system=None):
    """Procesa el dataframe agrupado por ECO y Año, devolviendo la comparativa limpia."""
    if filter_system:
        df_filtered = df_mes[df_mes['Sistema'] == filter_system].copy()
    else:
        df_filtered = df_mes.copy()
        
    if len(df_filtered) == 0:
        return pd.DataFrame(columns=['ECO', 'Sistemas_2025', 'Regs_2025', 'Importe_2025', 'Sistemas_2026', 'Regs_2026', 'Importe_2026', 'En_2025', 'En_2026', 'Presencia'])

    # 2025 Aggregation
    df_2025 = df_filtered[df_filtered['Anio'] == 2025].groupby('ECO').agg({
        'Sistema': lambda x: ', '.join(sorted(x.dropna().unique())),
        'Regs': 'sum',
        'Importe': 'sum'
    }).reset_index().rename(columns={
        'Sistema': 'Sistemas_2025',
        'Regs': 'Regs_2025',
        'Importe': 'Importe_2025'
    })

    # 2026 Aggregation
    df_2026 = df_filtered[df_filtered['Anio'] == 2026].groupby('ECO').agg({
        'Sistema': lambda x: ', '.join(sorted(x.dropna().unique())),
        'Regs': 'sum',
        'Importe': 'sum'
    }).reset_index().rename(columns={
        'Sistema': 'Sistemas_2026',
        'Regs': 'Regs_2026',
        'Importe': 'Importe_2026'
    })

    # Merge
    df_comp = pd.merge(df_2025, df_2026, on='ECO', how='outer').fillna({
        'Sistemas_2025': '',
        'Regs_2025': 0,
        'Importe_2025': 0.0,
        'Sistemas_2026': '',
        'Regs_2026': 0,
        'Importe_2026': 0.0
    })

    df_comp['En_2025'] = df_comp['Regs_2025'].apply(lambda r: 'SÍ' if r > 0 else 'NO')
    df_comp['En_2026'] = df_comp['Regs_2026'].apply(lambda r: 'SÍ' if r > 0 else 'NO')
    
    df_comp['Presencia'] = df_comp.apply(
        lambda r: 'AMBOS' if r['En_2025'] == 'SÍ' and r['En_2026'] == 'SÍ'
        else ('SOLO_2026' if r['En_2026'] == 'SÍ' else 'SOLO_2025'), axis=1
    )

    # Sort: AMBOS -> SOLO_2026 -> SOLO_2025
    df_comp['sort_cat'] = df_comp['Presencia'].map({'AMBOS': 1, 'SOLO_2026': 2, 'SOLO_2025': 3})
    df_comp = df_comp.sort_values(by=['sort_cat', 'ECO']).drop(columns=['sort_cat']).reset_index(drop=True)
    
    return df_comp

def main():
    parser = argparse.ArgumentParser(description="Generador de Reporte Comparativo de ECOs 2025 vs 2026 (Looker View)")
    parser.add_argument("--mes", type=int, default=1, help="Mes de comparación detallada (1-12, default: 1 = Enero)")
    args = parser.parse_args()

    mes = args.mes
    mes_nombre = get_month_name(mes)

    project_id = os.getenv("GCP_PROJECT_ID")
    dataset = os.getenv("BQ_DATASET", "rpa_utilitarios")
    view_name = os.getenv("BQ_VIEW_GASTOS", "vw_dashboard_gastos_final")
    view_id = f"{project_id}.{dataset}.{view_name}"

    if not project_id:
        print("❌ ERROR: GCP_PROJECT_ID no está en el .env")
        sys.exit(1)

    client = bigquery.Client(project=project_id)
    print(f"☁️ Conectando a la vista de BigQuery ({view_id})...")

    # 1. Consulta 1: Resumen General Mes por Mes (Vista de Looker)
    sql_general = f"""
        SELECT 
            EXTRACT(MONTH FROM Fecha) as Mes,
            COUNT(DISTINCT IF(EXTRACT(YEAR FROM Fecha) = 2025, ECO, NULL)) as Unidades_2025,
            COUNT(DISTINCT IF(EXTRACT(YEAR FROM Fecha) = 2026, ECO, NULL)) as Unidades_2026
        FROM `{view_id}`
        WHERE EXTRACT(YEAR FROM Fecha) IN (2025, 2026)
        GROUP BY 1
        ORDER BY 1
    """
    df_general = client.query(sql_general).to_dataframe()

    # 2. Consulta 2: Detalle del Mes seleccionado
    sql_detalle_mes = f"""
        SELECT 
            ECO,
            Sistema,
            EXTRACT(YEAR FROM Fecha) as Anio,
            COUNT(*) as Regs,
            SUM(Importe) as Importe
        FROM `{view_id}`
        WHERE EXTRACT(MONTH FROM Fecha) = {mes}
          AND EXTRACT(YEAR FROM Fecha) IN (2025, 2026)
        GROUP BY 1, 2, 3
        ORDER BY 1, 2, 3
    """
    df_mes = client.query(sql_detalle_mes).to_dataframe()

    wb = Workbook()
    wb.remove(wb.active) # Quitar hoja inicial

    # =========================================================================
    # HOJA 1: RESUMEN GENERAL (MES POR MES)
    # =========================================================================
    print("📈 Generando hoja RESUMEN GENERAL...")
    ws = wb.create_sheet(title="RESUMEN GENERAL")
    ws.sheet_view.showGridLines = True

    # Título
    ws.cell(row=2, column=2, value="CONCILIACIÓN GENERAL DE UNIDADES (2025 vs 2026)").font = FONT_TITLE
    ws.cell(row=3, column=2, value="Conteo de ECOs únicos activos en Looker (incluye mantenimientos y catálogo maestro)").font = Font(name=FONT_NAME, size=10, italic=True)

    headers = ["Mes", "Unidades 2025", "Unidades 2026", "Diferencia"]
    for idx, h in enumerate(headers, 2):
        cell = ws.cell(row=5, column=idx, value=h)
        format_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)

    # Llenar meses
    dict_general = df_general.set_index('Mes').to_dict('index')
    start_row = 6
    for m in range(1, 13):
        row = start_row + m - 1
        ws.cell(row=row, column=2, value=get_month_name(m))
        
        info = dict_general.get(m, {'Unidades_2025': 0, 'Unidades_2026': 0})
        ws.cell(row=row, column=3, value=int(info['Unidades_2025']))
        ws.cell(row=row, column=4, value=int(info['Unidades_2026']))
        
        # Fórmula de Excel para la diferencia
        ws.cell(row=row, column=5, value=f"=D{row}-C{row}")
        
        fill_row = FILL_ZEBRA if m % 2 == 0 else None
        
        format_cell(ws.cell(row=row, column=2), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws.cell(row=row, column=3), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws.cell(row=row, column=4), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        
        # Color en la diferencia
        diff = info['Unidades_2026'] - info['Unidades_2025']
        fill_diff = FILL_GREEN if diff > 0 else (FILL_RED if diff < 0 else fill_row)
        format_cell(ws.cell(row=row, column=5), font=FONT_BOLD, fill=fill_diff, alignment=ALIGN_CENTER)

    # Fila de Totales (Promedio de unidades)
    total_row = start_row + 12
    ws.cell(row=total_row, column=2, value="Promedio").font = FONT_BOLD
    ws.cell(row=total_row, column=3, value=f"=AVERAGE(C{start_row}:C{total_row-1})").font = FONT_BOLD
    ws.cell(row=total_row, column=4, value=f"=AVERAGE(D{start_row}:D{total_row-1})").font = FONT_BOLD
    ws.cell(row=total_row, column=5, value=f"=D{total_row}-C{total_row}").font = FONT_BOLD

    for c in range(2, 6):
        cell = ws.cell(row=total_row, column=c)
        cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
        if c in (3, 4, 5):
            cell.alignment = ALIGN_CENTER
            cell.number_format = '0.0'

    autofit_columns(ws, start_row=5)

    # =========================================================================
    # HOJA 2: COMPARATIVA GENERAL (MES SELECCIONADO)
    # =========================================================================
    print(f"📊 Generando hoja COMPARATIVA ({mes_nombre.upper()})...")
    ws_comp = wb.create_sheet(title=f"COMPARATIVA {mes_nombre[:3].upper()}")
    ws_comp.sheet_view.showGridLines = True

    ws_comp.cell(row=2, column=2, value=f"COMPARATIVA GENERAL DE ECOs — {mes_nombre.upper()}").font = FONT_TITLE
    ws_comp.cell(row=3, column=2, value=f"Detalle de ECOs activos en Looker en {mes_nombre} 2025 vs 2026").font = Font(name=FONT_NAME, size=10, italic=True)

    headers_comp = ["ECO", "En 2025?", "En 2026?", "Sistemas 2025", "Sistemas 2026", "Regs 2025", "Importe 2025", "Regs 2026", "Importe 2026", "Presencia"]
    for idx, h in enumerate(headers_comp, 2):
        cell = ws_comp.cell(row=5, column=idx, value=h)
        format_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)

    df_comp_data = generate_comparative_df(df_mes)
    for idx, row in df_comp_data.iterrows():
        r_num = 6 + idx
        fill_row = FILL_ZEBRA if idx % 2 == 0 else None
        
        # Color basado en Presencia
        fill_presencia = fill_row
        if row['Presencia'] == 'SOLO_2026':
            fill_presencia = FILL_GREEN
        elif row['Presencia'] == 'SOLO_2025':
            fill_presencia = FILL_RED

        ws_comp.cell(row=r_num, column=2, value=row['ECO'])
        ws_comp.cell(row=r_num, column=3, value=row['En_2025'])
        ws_comp.cell(row=r_num, column=4, value=row['En_2026'])
        ws_comp.cell(row=r_num, column=5, value=row['Sistemas_2025'])
        ws_comp.cell(row=r_num, column=6, value=row['Sistemas_2026'])
        ws_comp.cell(row=r_num, column=7, value=int(row['Regs_2025']))
        ws_comp.cell(row=r_num, column=8, value=float(row['Importe_2025']))
        ws_comp.cell(row=r_num, column=9, value=int(row['Regs_2026']))
        ws_comp.cell(row=r_num, column=10, value=float(row['Importe_2026']))
        ws_comp.cell(row=r_num, column=11, value=row['Presencia'])

        format_cell(ws_comp.cell(row=r_num, column=2), font=FONT_BOLD, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_comp.cell(row=r_num, column=3), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws_comp.cell(row=r_num, column=4), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws_comp.cell(row=r_num, column=5), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_comp.cell(row=r_num, column=6), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_comp.cell(row=r_num, column=7), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws_comp.cell(row=r_num, column=8), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_RIGHT, num_format='$#,##0.00')
        format_cell(ws_comp.cell(row=r_num, column=9), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws_comp.cell(row=r_num, column=10), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_RIGHT, num_format='$#,##0.00')
        format_cell(ws_comp.cell(row=r_num, column=11), font=FONT_BOLD, fill=fill_presencia, alignment=ALIGN_CENTER)

    # Totales comparativa general
    tot_row = 6 + len(df_comp_data)
    ws_comp.cell(row=tot_row, column=2, value="Total").font = FONT_BOLD
    ws_comp.cell(row=tot_row, column=7, value=f"=SUM(G6:G{tot_row-1})").font = FONT_BOLD
    ws_comp.cell(row=tot_row, column=8, value=f"=SUM(H6:H{tot_row-1})").font = FONT_BOLD
    ws_comp.cell(row=tot_row, column=9, value=f"=SUM(I6:I{tot_row-1})").font = FONT_BOLD
    ws_comp.cell(row=tot_row, column=10, value=f"=SUM(J6:J{tot_row-1})").font = FONT_BOLD

    for c in range(2, 12):
        cell = ws_comp.cell(row=tot_row, column=c)
        cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
        if c in (7, 9):
            cell.alignment = ALIGN_CENTER
        elif c in (8, 10):
            cell.alignment = ALIGN_RIGHT
            cell.number_format = '$#,##0.00'

    autofit_columns(ws_comp, start_row=5)


    # =========================================================================
    # HOJAS 3, 4, 5, 6: SISTEMAS (PASE, SUPRAMAX, EDENRED, MANTENIMIENTOS)
    # =========================================================================
    sistemas_config = [
        ("Pase", "PASE"),
        ("Supramax", "SUPRAMAX"),
        ("Edenred", "EDENRED"),
        ("Google Sheets", "MANTENIMIENTOS")
    ]
    
    for sys_db_name, sheet_name in sistemas_config:
        print(f"🛣️/💳/🔧 Generando hoja {sheet_name}...")
        ws_sys = wb.create_sheet(title=sheet_name)
        ws_sys.sheet_view.showGridLines = True

        ws_sys.cell(row=2, column=2, value=f"COMPARATIVA ECOs — {sheet_name} ({mes_nombre.upper()})").font = FONT_TITLE
        ws_sys.cell(row=3, column=2, value=f"Detalle exclusivo de transacciones activas en {sheet_name} ({sys_db_name})").font = Font(name=FONT_NAME, size=10, italic=True)

        headers_sys = ["ECO", "Regs 2025", "Importe 2025", "Regs 2026", "Importe 2026", "Presencia"]
        for idx, h in enumerate(headers_sys, 2):
            cell = ws_sys.cell(row=5, column=idx, value=h)
            format_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)

        df_sys_data = generate_comparative_df(df_mes, filter_system=sys_db_name)
        
        for idx, row in df_sys_data.iterrows():
            r_num = 6 + idx
            fill_row = FILL_ZEBRA if idx % 2 == 0 else None
            
            fill_presencia = fill_row
            if row['Presencia'] == 'SOLO_2026':
                fill_presencia = FILL_GREEN
            elif row['Presencia'] == 'SOLO_2025':
                fill_presencia = FILL_RED

            ws_sys.cell(row=r_num, column=2, value=row['ECO'])
            ws_sys.cell(row=r_num, column=3, value=int(row['Regs_2025']))
            ws_sys.cell(row=r_num, column=4, value=float(row['Importe_2025']))
            ws_sys.cell(row=r_num, column=5, value=int(row['Regs_2026']))
            ws_sys.cell(row=r_num, column=6, value=float(row['Importe_2026']))
            ws_sys.cell(row=r_num, column=7, value=row['Presencia'])

            format_cell(ws_sys.cell(row=r_num, column=2), font=FONT_BOLD, fill=fill_row, alignment=ALIGN_LEFT)
            format_cell(ws_sys.cell(row=r_num, column=3), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
            format_cell(ws_sys.cell(row=r_num, column=4), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_RIGHT, num_format='$#,##0.00')
            format_cell(ws_sys.cell(row=r_num, column=5), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
            format_cell(ws_sys.cell(row=r_num, column=6), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_RIGHT, num_format='$#,##0.00')
            format_cell(ws_sys.cell(row=r_num, column=7), font=FONT_BOLD, fill=fill_presencia, alignment=ALIGN_CENTER)

        # Totales por sistema
        tot_row_sys = 6 + len(df_sys_data)
        ws_sys.cell(row=tot_row_sys, column=2, value="Total").font = FONT_BOLD
        ws_sys.cell(row=tot_row_sys, column=3, value=f"=SUM(C6:C{tot_row_sys-1})").font = FONT_BOLD
        ws_sys.cell(row=tot_row_sys, column=4, value=f"=SUM(D6:D{tot_row_sys-1})").font = FONT_BOLD
        ws_sys.cell(row=tot_row_sys, column=5, value=f"=SUM(E6:E{tot_row_sys-1})").font = FONT_BOLD
        ws_sys.cell(row=tot_row_sys, column=6, value=f"=SUM(F6:F{tot_row_sys-1})").font = FONT_BOLD

        for c in range(2, 8):
            cell = ws_sys.cell(row=tot_row_sys, column=c)
            cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
            if c in (3, 5):
                cell.alignment = ALIGN_CENTER
            elif c in (4, 6):
                cell.alignment = ALIGN_RIGHT
                cell.number_format = '$#,##0.00'

        autofit_columns(ws_sys, start_row=5)

    # =========================================================================
    # HOJA 7: VALIDACION CATALOGO (NUEVO)
    # =========================================================================
    print("📋 Generando hoja VALIDACION CATALOGO...")
    ws_cat = wb.create_sheet(title="VALIDACION CATALOGO")
    ws_cat.sheet_view.showGridLines = True

    ws_cat.cell(row=2, column=2, value="VALIDACIÓN DE CATÁLOGO MAESTRO (TODO EL HISTÓRICO)").font = FONT_TITLE
    ws_cat.cell(row=3, column=2, value="Lista completa de ECOs en transacciones, indicando si existen en la Tabla Maestra y sus nombres originales").font = Font(name=FONT_NAME, size=10, italic=True)

    headers_cat = [
        "ECO (Normalizado)", "ECO Original (en Transacciones)", "¿En Tabla Maestra?", 
        "Empresa (Catálogo)", "Empresa (Transacciones)", "Sistemas", 
        "Transacciones Totales", "Importe Total", "Primera Transacción", "Última Transacción"
    ]
    for idx, h in enumerate(headers_cat, 2):
        cell = ws_cat.cell(row=5, column=idx, value=h)
        format_cell(cell, font=FONT_HEADER, fill=FILL_HEADER, alignment=ALIGN_CENTER)

    # Consulta a BigQuery
    sql_cat = f"""
        SELECT 
            COALESCE(m.ECO, g.ECO) as ECO,
            IF(m.ECO IS NULL, 'NO', 'SÍ') as En_Tabla_Maestra,
            IFNULL(m.EMPRESA, 'N/A') as Empresa_Catalogo,
            IFNULL(STRING_AGG(DISTINCT g.Sistema), 'SIN ACTIVIDAD') as Sistemas,
            IFNULL(STRING_AGG(DISTINCT IFNULL(g.Empresa, 'N/A')), 'N/A') as Empresas_en_Txs,
            COUNT(g.ECO) as Total_Transacciones,
            ROUND(IFNULL(SUM(g.Importe), 0.0), 2) as Total_Importe,
            MIN(g.Fecha) as Primera_Transaccion,
            MAX(g.Fecha) as Ultima_Transaccion
        FROM `{project_id}.{dataset}.tbl_utilitarios_maestra` m
        FULL OUTER JOIN `{project_id}.{dataset}.consumos_flota` g ON m.ECO = g.ECO
        GROUP BY 1, 2, 3
        ORDER BY En_Tabla_Maestra, ECO
    """
    df_cat_data = client.query(sql_cat).to_dataframe()

    # Mapeo de ECOs originales (LZC)
    orig_map = get_original_eco_map()

    for idx, row in df_cat_data.iterrows():
        r_num = 6 + idx
        fill_row = FILL_ZEBRA if idx % 2 == 0 else None
        
        # Color basado en si está o no en la tabla maestra
        if row['En_Tabla_Maestra'] == 'NO':
            fill_status = FILL_RED
        else:
            fill_status = FILL_GREEN

        eco_norm = row['ECO']
        eco_orig = orig_map.get(eco_norm, eco_norm)

        # Formatear fechas como cadenas limpias
        f_min = str(row['Primera_Transaccion']) if pd.notna(row['Primera_Transaccion']) else ''
        f_max = str(row['Ultima_Transaccion']) if pd.notna(row['Ultima_Transaccion']) else ''

        ws_cat.cell(row=r_num, column=2, value=eco_norm)
        ws_cat.cell(row=r_num, column=3, value=eco_orig)
        ws_cat.cell(row=r_num, column=4, value=row['En_Tabla_Maestra'])
        ws_cat.cell(row=r_num, column=5, value=row['Empresa_Catalogo'])
        ws_cat.cell(row=r_num, column=6, value=row['Empresas_en_Txs'])
        ws_cat.cell(row=r_num, column=7, value=row['Sistemas'])
        ws_cat.cell(row=r_num, column=8, value=int(row['Total_Transacciones']) if pd.notna(row['Total_Transacciones']) else 0)
        ws_cat.cell(row=r_num, column=9, value=float(row['Total_Importe']) if pd.notna(row['Total_Importe']) else 0.0)
        ws_cat.cell(row=r_num, column=10, value=f_min)
        ws_cat.cell(row=r_num, column=11, value=f_max)

        format_cell(ws_cat.cell(row=r_num, column=2), font=FONT_BOLD, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_cat.cell(row=r_num, column=3), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_cat.cell(row=r_num, column=4), font=FONT_BOLD, fill=fill_status, alignment=ALIGN_CENTER)
        format_cell(ws_cat.cell(row=r_num, column=5), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_cat.cell(row=r_num, column=6), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_cat.cell(row=r_num, column=7), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_LEFT)
        format_cell(ws_cat.cell(row=r_num, column=8), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws_cat.cell(row=r_num, column=9), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_RIGHT, num_format='$#,##0.00')
        format_cell(ws_cat.cell(row=r_num, column=10), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)
        format_cell(ws_cat.cell(row=r_num, column=11), font=FONT_NORMAL, fill=fill_row, alignment=ALIGN_CENTER)

    # Fila de totales
    tot_row_cat = 6 + len(df_cat_data)
    ws_cat.cell(row=tot_row_cat, column=2, value="Total").font = FONT_BOLD
    ws_cat.cell(row=tot_row_cat, column=8, value=f"=SUM(H6:H{tot_row_cat-1})").font = FONT_BOLD
    ws_cat.cell(row=tot_row_cat, column=9, value=f"=SUM(I6:I{tot_row_cat-1})").font = FONT_BOLD

    for c in range(2, 12):
        cell = ws_cat.cell(row=tot_row_cat, column=c)
        cell.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
        if c == 8:
            cell.alignment = ALIGN_CENTER
        elif c == 9:
            cell.alignment = ALIGN_RIGHT
            cell.number_format = '$#,##0.00'

    autofit_columns(ws_cat, start_row=5)

    out_file = "comparativa_ecos_2025_vs_2026.xlsx"
    print(f"\n💾 Guardando reporte en: {out_file}...")
    wb.save(out_file)
    print("🎉 Reporte guardado con éxito.")

if __name__ == "__main__":
    main()
