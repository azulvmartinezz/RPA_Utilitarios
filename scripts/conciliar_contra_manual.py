import os
import re
import sys
import pandas as pd
from google.cloud import bigquery
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

_BQ_IDENTIFIER_RE = re.compile(r'^[A-Za-z0-9_-]+$')


def _safe_bq_identifier(value, name):
    if not value or not _BQ_IDENTIFIER_RE.match(str(value)):
        raise ValueError(f"{name} contiene caracteres no permitidos")
    return value

def get_month_name(ym):
    meses = {'01':'Enero', '02':'Febrero', '03':'Marzo', '04':'Abril', '05':'Mayo', '06':'Junio', 
             '07':'Julio', '08':'Agosto', '09':'Septiembre', '10':'Octubre', '11':'Noviembre', '12':'Diciembre'}
    parts = str(ym).split('-')
    if len(parts) == 2:
        return f"{meses.get(parts[1], parts[1])} {parts[0]}"
    return ym

def _limpiar_eco(s):
    return str(s).strip().upper().replace(' ', '')

def main():
    print("Iniciando Conciliación Automática...")
    load_dotenv()
    
    # --- CONFIGURACIÓN DE ARCHIVOS (Vía .env para seguridad) ---
    manual_file = os.getenv('FILE_GASTOS_MANUAL', 'MES POR MES - GASTOS 2026.xlsx')

    if not os.path.exists(manual_file):
        print(f"❌ No se encontró el archivo de gastos manuales: {manual_file}")
        print("Asegúrate de definir FILE_GASTOS_MANUAL en tu .env")
        return
        
    df_manual = pd.read_excel(manual_file, sheet_name=0)
    df_manual.columns = df_manual.columns.str.strip()
    
    df_manual['ECO_Limpio'] = df_manual['UNIDAD'].apply(_limpiar_eco)
    df_manual['MES_str'] = pd.to_datetime(df_manual['MES'], errors='coerce').dt.strftime('%Y-%m')
    
    manual_sums = df_manual.groupby(['ECO_Limpio', 'MES_str']).agg({
        'IAVE PASE': 'sum',
        'SUPRAMAX': 'sum',
        'TICKET CARD': 'sum',
        'MANTENIMIENTO': 'sum'
    }).fillna(0).reset_index()

    project_id = os.getenv('GCP_PROJECT_ID')
    dataset = os.getenv('BQ_DATASET')

    if not project_id or not dataset:
        print("❌ ERROR: No se encontró GCP_PROJECT_ID o BQ_DATASET en el .env")
        return
        
    # Usamos la vista final que ya consolida todo (incluyendo Google Sheets)
    project_id = _safe_bq_identifier(project_id, 'GCP_PROJECT_ID')
    dataset = _safe_bq_identifier(dataset, 'BQ_DATASET')
    view_name = _safe_bq_identifier(os.getenv('BQ_VIEW_GASTOS', 'vw_dashboard_gastos_final'), 'BQ_VIEW_GASTOS')
    view_id = f"{project_id}.{dataset}.{view_name}"
    
    client = bigquery.Client(project=project_id)
    query = f"""
        SELECT 
            ECO as ECO_Limpio,
            FORMAT_DATE('%Y-%m', DATE_TRUNC(Fecha, MONTH)) as MES_str,
            Sistema,
            SUM(Importe) as Total_BQ
        FROM `{view_id}`
        WHERE Fecha >= '2026-01-01'
        GROUP BY ECO, MES_str, Sistema
    """
    df_bq = client.query(query).to_dataframe()
    
    df_bq_pivot = df_bq.pivot_table(index=['ECO_Limpio', 'MES_str'], columns='Sistema', values='Total_BQ', fill_value=0).reset_index()
    
    # Asegurar que todas las columnas existan
    for sis in ['Pase', 'Supramax', 'Edenred', 'Google Sheets']:
        if sis not in df_bq_pivot.columns: df_bq_pivot[sis] = 0.0
            
    # Unir todo (Manual vs BigQuery)
    df_cruce = pd.merge(manual_sums, df_bq_pivot, on=['ECO_Limpio', 'MES_str'], how='outer').fillna(0)
    
    # Mapeo de sistemas: (Nombre Hoja, Columna Manual, Columna BQ)
    sistemas = [
        ('Pase', 'IAVE PASE', 'Pase'),
        ('Supramax', 'SUPRAMAX', 'Supramax'),
        ('Edenred', 'TICKET CARD', 'Edenred'),
        ('Mantenimientos', 'MANTENIMIENTO', 'Google Sheets')
    ]

    # ====== GENERAR EXCEL ESTÉTICO ======
    wb = Workbook()
    wb.remove(wb.active) # Quitar hoja por defecto
    
    # Fuentes y estilos globales
    century_font = Font(name='Century Gothic', size=11)
    century_bold = Font(name='Century Gothic', size=11, bold=True)
    century_title = Font(name='Century Gothic', size=14, bold=True)
    align_left = Alignment(horizontal='left', vertical='center')
    
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name='Century Gothic', bold=True, color="FFFFFF", size=11)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # CREAR PESTAÑA DE RESUMEN
    ws_resumen = wb.create_sheet(title="RESUMEN")
    ws_resumen.sheet_view.showGridLines = False
    row_idx = 2
    
    meses_unicos = sorted(df_cruce['MES_str'].dropna().unique())
    
    for mes in meses_unicos:
        mes_bonito = get_month_name(mes).upper()
        
        # Título del Mes Grande
        cell_mes = ws_resumen.cell(row=row_idx, column=2, value=f"      MES: {mes_bonito}      ")
        cell_mes.font = Font(name='Century Gothic', size=16, bold=True, color="1F4E78")
        row_idx += 2
        
        for nombre_hoja, col_manual, col_bq in sistemas:
            df_sis = df_cruce[(df_cruce['MES_str'] == mes) & ((df_cruce[col_manual] > 0) | (df_cruce[col_bq] > 0))].copy()
            
            tot_sis = df_sis[col_manual].sum()
            tot_cons = df_sis[col_bq].sum()
            
            # Si no hay datos ni manuales ni en BQ para este sistema en este mes, saltarlo
            if tot_sis == 0 and tot_cons == 0:
                continue
                
            dif_tot = tot_cons - tot_sis
            var_pct = (dif_tot / tot_sis) if tot_sis != 0 else 0
            
            solo_cons = df_sis[(df_sis[col_manual] == 0) & (df_sis[col_bq] > 0)]
            solo_sis = df_sis[(df_sis[col_manual] > 0) & (df_sis[col_bq] == 0)]
            ambos = df_sis[(df_sis[col_manual] > 0) & (df_sis[col_bq] > 0)]
            
            # Título Sistema
            ws_resumen.cell(row=row_idx, column=2, value=f"RESUMEN {nombre_hoja.upper()}").font = century_title
            row_idx += 2
            
            # Bloque RESUMEN
            ws_resumen.cell(row=row_idx, column=2, value="RESUMEN").font = century_bold
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value="Total Reporte Manual (Depto):")
            ws_resumen.cell(row=row_idx, column=3, value=tot_sis).number_format = '$#,##0.00'
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value="Total Automatizado (BigQuery):")
            ws_resumen.cell(row=row_idx, column=3, value=tot_cons).number_format = '$#,##0.00'
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value="Diferencia Total:")
            ws_resumen.cell(row=row_idx, column=3, value=dif_tot).number_format = '$#,##0.00'
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value="Variación %:")
            ws_resumen.cell(row=row_idx, column=3, value=var_pct).number_format = '0.0%'
            row_idx += 2
            
            # Bloque ANÁLISIS
            ws_resumen.cell(row=row_idx, column=2, value="ANÁLISIS").font = century_bold
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value=f"Unidades solo en Automático (Faltan en Manual): {len(solo_cons)}")
            ws_resumen.cell(row=row_idx, column=3, value=solo_cons[col_bq].sum()).number_format = '$#,##0.00'
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value=f"Unidades solo en Manual (Faltan en Automático): {len(solo_sis)}")
            ws_resumen.cell(row=row_idx, column=3, value=solo_sis[col_manual].sum()).number_format = '$#,##0.00'
            row_idx += 1
            ws_resumen.cell(row=row_idx, column=2, value=f"Unidades reportadas por ambos: {len(ambos)}")
            ws_resumen.cell(row=row_idx, column=3, value=len(ambos))
            row_idx += 4

    # Aplicar formato a toda la hoja RESUMEN
    for r in range(1, row_idx):
        if ws_resumen.row_dimensions[r].height is None:
            ws_resumen.row_dimensions[r].height = 20
        for c in range(1, 10): # Extendemos a más columnas por la cuadrícula
            cell = ws_resumen.cell(row=r, column=c)
            if not cell.alignment.wrap_text: # No sobreescribir el wrap_text de las listas de ECOs
                cell.alignment = align_left
            if not cell.font or cell.font.name != 'Century Gothic':
                is_bold = cell.font.bold if cell.font else False
                sz = cell.font.size if cell.font and cell.font.size else 11
                it = cell.font.italic if cell.font else False
                color = cell.font.color if cell.font else "000000"
                cell.font = Font(name='Century Gothic', size=sz, bold=is_bold, italic=it, color=color)

    ws_resumen.column_dimensions['A'].width = 5
    ws_resumen.column_dimensions['B'].width = 45
    ws_resumen.column_dimensions['C'].width = 25

    # CREAR PESTAÑAS POR SISTEMA
    for nombre_hoja, col_manual, col_bq in sistemas:
        ws = wb.create_sheet(title=nombre_hoja)
        ws.sheet_view.showGridLines = False
        
        headers = ['Mes', 'Unidad', f'{nombre_hoja} (Manual)', f'{nombre_hoja} (BigQuery)', 'Diferencia', 'Estado']
        ws.append(headers)
        ws.row_dimensions[1].height = 20
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_left
            cell.border = border
            
        df_sis = df_cruce[(df_cruce[col_manual] > 0) | (df_cruce[col_bq] > 0)].copy()
        df_sis = df_sis.sort_values(['MES_str', 'ECO_Limpio'])
        
        for idx, row in df_sis.iterrows():
            mes_bonito = get_month_name(row['MES_str'])
            unit = row['ECO_Limpio']
            sist = round(row[col_manual], 2)
            bq = round(row[col_bq], 2)
            dif = round(bq - sist, 2)
            
            if sist == 0 and bq > 0:
                status = "❌ FALTA EN REPORTE MANUAL"
                fill_color = "FFE699"
            elif sist > 0 and bq == 0:
                status = "⚠️ FALTA EN BOT AUTOMÁTICO"
                fill_color = "F4B084"
            elif sist > 0 and bq > 0 and dif != 0:
                status = "📊 DIFERENCIA DE MONTO"
                fill_color = "C5D9F1"
            elif sist > 0 and bq > 0:
                status = "✅ CUADRA PERFECTO"
                fill_color = "C6EFCE"
            else:
                status = "⚪ SIN MOVIMIENTO"
                fill_color = "D9D9D9"
                
            ws.append([mes_bonito, unit, sist, bq, dif, status])
            
            row_num = ws.max_row
            ws.row_dimensions[row_num].height = 20
            
            for col_num, cell in enumerate(ws[row_num], 1):
                cell.font = century_font
                cell.border = border
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.alignment = align_left
                if col_num in [3, 4, 5]:
                    cell.number_format = '$#,##0.00'
                    
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 45
        
        # Activar AutoFiltros
        ws.auto_filter.ref = ws.dimensions

    wb.save('REPORTE_CONCILIACION_FLOTA.xlsx')
    print("✅ Archivo creado con RESUMEN: REPORTE_CONCILIACION_FLOTA.xlsx")

if __name__ == "__main__":
    main()
