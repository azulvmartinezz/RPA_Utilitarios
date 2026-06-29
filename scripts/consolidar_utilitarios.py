import os
import re
import sys
import warnings
import unicodedata
import pandas as pd
from dotenv import load_dotenv

# Añadir el directorio raíz al path para importaciones
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if PROJECT_ROOT not in sys.path:
    sys.path.insert(0, PROJECT_ROOT)

load_dotenv()

def _normalize_eco(val):
    s = str(val).strip().upper().replace('.', '')
    # Check if it has LZC
    has_lzc = 'LZC' in s
    s_clean = s.replace('LZC', '').replace(' ', '')
    
    m = re.match(r'^(AU|CA)-?(\d{1,3})(?!\d)', s_clean)
    if m:
        eco = f"{m.group(1)}-{m.group(2).zfill(3)}"
        if has_lzc:
            return f"{eco} LZC"
        return eco
    return str(val).strip().upper()

def _clean_eco_key(val):
    return re.sub(r'[^A-Z0-9]', '', str(val).upper())


def _normalize_col_name(name):
    text = unicodedata.normalize("NFKD", str(name)).encode("ascii", "ignore").decode("ascii")
    return re.sub(r'[^a-z0-9]+', '', text.lower())


def _expand_tabbed_single_column(df):
    real_cols = [c for c in df.columns if not str(c).startswith("Unnamed:")]
    if len(real_cols) != 1:
        return df

    first_col = real_cols[0]
    if "\t" not in str(first_col):
        return df

    header_parts = [part.strip() for part in str(first_col).split("\t")]
    rows = []
    for value in df[first_col].fillna(""):
        parts = [part.strip() for part in str(value).split("\t")]
        if len(parts) < len(header_parts):
            parts.extend([""] * (len(header_parts) - len(parts)))
        rows.append(parts[:len(header_parts)])
    return pd.DataFrame(rows, columns=header_parts)


def _find_eco_column(columns):
    normalized = {col: _normalize_col_name(col) for col in columns}
    for col, norm in normalized.items():
        if norm == "eco":
            return col
    for col, norm in normalized.items():
        if norm in {"noeconomico", "numeroeconomico"} or ("economico" in norm and "no" in norm):
            return col
    return None


def _load_maestra_dataframe(maestro_path):
    explicit_sheet_name = os.getenv("EXCEL_MAESTRO_SHEET")
    preferred_sheet_names = [
        "Datos_Asignación",
        "Datos_Asignacion",
        "Maestra_Consolidada",
        "_Join",
        "Datos_Unidad",
        "Sheet1",
    ]
    if explicit_sheet_name:
        preferred_sheet_names.insert(0, explicit_sheet_name)

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Data Validation extension is not supported and will be removed",
            category=UserWarning,
        )
        excel_file = pd.ExcelFile(maestro_path)

    candidate_sheets = []
    seen = set()
    for sheet_name in preferred_sheet_names + excel_file.sheet_names:
        if sheet_name not in seen:
            seen.add(sheet_name)
            candidate_sheets.append(sheet_name)

    inspected = []
    valid_candidates = []
    for sheet_name in candidate_sheets:
        try:
            with warnings.catch_warnings():
                warnings.filterwarnings(
                    "ignore",
                    message="Data Validation extension is not supported and will be removed",
                    category=UserWarning,
                )
                warnings.filterwarnings(
                    "ignore",
                    message="Conditional Formatting extension is not supported and will be removed",
                    category=UserWarning,
                )
                df_sheet = pd.read_excel(maestro_path, sheet_name=sheet_name)
        except Exception:
            continue

        df_sheet = _expand_tabbed_single_column(df_sheet)
        df_sheet.columns = [str(col).strip() for col in df_sheet.columns]
        eco_col = _find_eco_column(df_sheet.columns)
        inspected.append((sheet_name, eco_col, df_sheet.shape))
        if not eco_col:
            continue

        rows_with_eco = df_sheet[eco_col].astype(str).str.strip().ne("").sum()
        valid_candidates.append(
            {
                "sheet_name": sheet_name,
                "eco_col": eco_col,
                "df": df_sheet,
                "score": (rows_with_eco, len(df_sheet.columns)),
            }
        )

    if not valid_candidates:
        detail = ", ".join(
            f"{sheet}={shape}, eco_col={eco_col}"
            for sheet, eco_col, shape in inspected
        )
        raise KeyError(
            "No se encontró una hoja válida con columna ECO en la Tabla Maestra. "
            f"Hojas inspeccionadas: {detail}"
        )

    best = None
    for preferred_name in preferred_sheet_names:
        best = next((item for item in valid_candidates if item["sheet_name"] == preferred_name), None)
        if best is not None:
            break
    if best is None:
        best = max(valid_candidates, key=lambda item: item["score"])

    df_maestra = best["df"].copy()
    if best["eco_col"] != "ECO":
        df_maestra = df_maestra.rename(columns={best["eco_col"]: "ECO"})
    print(
        f"✅ Hoja de Tabla Maestra seleccionada automáticamente: "
        f"{best['sheet_name']} ({len(df_maestra)} filas)."
    )
    return df_maestra


def _find_first_column(columns, *candidates):
    normalized = {col: _normalize_col_name(col) for col in columns}
    for candidate in candidates:
        target = _normalize_col_name(candidate)
        for col, norm in normalized.items():
            if norm == target:
                return col
    return None

def consolidar_todo():
    print("=== INICIANDO PROCESO DE CONSOLIDACIÓN LOCAL ===")
    
    # 1. Obtener rutas desde el archivo .env
    maestro_path = os.getenv('EXCEL_MAESTRO_PATH')
    mantenimiento_path = os.getenv('EXCEL_MANTENIMIENTO_PATH')
    
    # Validar que los archivos existan
    if not maestro_path or not os.path.exists(maestro_path):
        print(f"❌ Error: No se encontró el archivo de Tabla Maestra en la ruta: {maestro_path}")
        return
        
    if not mantenimiento_path or not os.path.exists(mantenimiento_path):
        print(f"⚠️ Advertencia: No se encontró el archivo de Mantenimientos en la ruta: {mantenimiento_path}. Se procederá sin mantenimientos.")
        df_mantenimientos_raw = pd.DataFrame()
    else:
        try:
            # Leer la pestaña BASE del Excel de Mantenimientos
            df_mantenimientos_raw = pd.read_excel(mantenimiento_path, sheet_name='BASE')
            print(f"✅ Archivo de Mantenimientos (pestaña BASE) cargado con éxito ({len(df_mantenimientos_raw)} filas).")
        except PermissionError:
            print(f"❌ Error de permisos: El archivo de Mantenimientos '{mantenimiento_path}' está siendo usado por otro programa (probablemente está abierto en Excel). Por favor, ciérralo e intenta de nuevo.")
            df_mantenimientos_raw = pd.DataFrame()
        except Exception as e:
            print(f"❌ Error al leer la pestaña BASE del archivo de Mantenimientos: {e}")
            df_mantenimientos_raw = pd.DataFrame()

    try:
        # Cargar Tabla Maestra detectando automáticamente la hoja tabular correcta.
        df_maestra = _load_maestra_dataframe(maestro_path)
        print(f"✅ Archivo de Tabla Maestra cargado con éxito ({len(df_maestra)} filas).")
    except PermissionError:
        print(f"❌ Error de permisos: El archivo de Tabla Maestra '{maestro_path}' está siendo usado por otro programa (probablemente está abierto en Excel). Por favor, ciérralo e intenta de nuevo.")
        return
    except Exception as e:
        print(f"❌ Error crítico al leer el archivo de Tabla Maestra: {e}")
        return

    # Construir diccionario de mapeo de Placas/Supramax ID a ECO
    mapping_eco = {}
    for _, row in df_maestra.iterrows():
        eco = str(row.get('ECO', '')).strip().upper()
        if re.match(r'^(AU|CA)-?\d{3}(?:\s*LZC)?$', eco):
            eco_clean = eco.replace('AU', 'AU-').replace('CA', 'CA-').replace('--', '-')
            # Ensure proper spacing for LZC
            if 'LZC' in eco_clean and ' LZC' not in eco_clean:
                eco_clean = eco_clean.replace('LZC', ' LZC')
            eco = eco_clean
            
            # Mapear Placas
            placa = str(row.get('Placas', '')).strip().upper().replace(' ', '').replace('-', '')
            if placa and placa != 'NAN':
                mapping_eco[placa] = eco
            # Mapear Supramax ID
            supra = str(row.get('Supramax ID', '')).strip().upper()
            if supra and supra != 'NAN':
                mapping_eco[supra] = eco

    # Función auxiliar para mapear ECOs
    def _resolve_eco(val):
        val_norm = _normalize_eco(val)
        if re.match(r'^(AU|CA)-\d{3}(?:\s*LZC)?$', val_norm):
            return val_norm
        # Intentar buscar como placa
        val_placa = str(val).strip().upper().replace(' ', '').replace('-', '')
        if val_placa in mapping_eco:
            return mapping_eco[val_placa]
        # Intentar buscar como Supramax ID exacto
        val_supra = str(val).strip().upper()
        if val_supra in mapping_eco:
            return mapping_eco[val_supra]
        return val_norm

    # 2. Cargar Consumos Consolidados (Pase, Supramax, Edenred)
    lista_consumos = []
    
    for sistema, archivo in [('Pase', 'CONSOLIDADO_CRUDO_PASE.csv'), 
                             ('Supramax', 'CONSOLIDADO_CRUDO_SUPRAMAX.csv'), 
                             ('Edenred', 'CONSOLIDADO_LIMPIO_EDENRED.csv')]:
        if os.path.exists(archivo):
            try:
                header_columns = pd.read_csv(archivo, nrows=0).columns.tolist()
                col_eco = _find_first_column(header_columns, 'ECO', 'PLACAS', 'No. economico', 'No Economico')
                col_fecha = _find_first_column(header_columns, 'Fecha', 'FECHA', 'Fecha Transacción')
                col_importe = _find_first_column(header_columns, 'Importe', 'IMPORTE', 'Importe Transacción')
                col_cantidad = _find_first_column(header_columns, 'Cantidad', 'CANTIDAD', 'Cantidad Mercancía')
                col_tipo = _find_first_column(header_columns, 'Tipo', 'PRODUCTO', 'Mercancía')
                col_concepto = _find_first_column(header_columns, 'Concepto')
                col_tarjeta = _find_first_column(header_columns, 'Tarjeta IDMX')

                if not col_eco or not col_fecha or not col_importe:
                    raise KeyError(
                        f"Columnas requeridas no encontradas en {archivo}. "
                        f"ECO={col_eco}, Fecha={col_fecha}, Importe={col_importe}. "
                        f"Disponibles: {header_columns}"
                    )

                selected_columns = [col_eco, col_fecha, col_importe, col_cantidad, col_tipo, col_concepto, col_tarjeta]
                usecols = list(dict.fromkeys([col for col in selected_columns if col]))
                df_c = pd.read_csv(archivo, usecols=usecols)
                df_c['Sistema'] = sistema
                # Normalizar nombres de columnas a minúsculas para búsqueda flexible
                cols_lower = {c.lower(): c for c in df_c.columns}
                
                # Detectar columna ECO/PLACAS
                col_eco = None
                for kw in ['eco', 'placas', 'placa']:
                    if kw in cols_lower:
                        col_eco = cols_lower[kw]
                        break
                
                # Detectar columna Fecha
                col_fecha = None
                for kw in ['fecha', 'datetime', 'date']:
                    if kw in cols_lower:
                        col_fecha = cols_lower[kw]
                        break
                
                # Detectar columna Importe
                col_importe = None
                for kw in ['importe', 'total', 'monto', 'subtotal']:
                    if kw in cols_lower:
                        col_importe = cols_lower[kw]
                        break
                
                if not col_eco or not col_fecha or not col_importe:
                    print(f"⚠️ Columnas requeridas no detectadas en {archivo}. Se necesitan equivalentes a ECO, Fecha e Importe.")
                    continue
                
                # Estandarizar columnas
                df_c['__fecha_std'] = pd.to_datetime(df_c[col_fecha], errors='coerce')
                
                # Mapeo de columnas específicas
                if col_tarjeta:
                    df_c['Concepto'] = 'PEAJE'
                    df_c['Tipo'] = None
                    df_c['Cantidad'] = None
                elif sistema == 'Supramax':
                    # Si no tiene concepto, asignamos combustible
                    if not col_concepto:
                        df_c['Concepto'] = 'COMBUSTIBLE'
                
                df_std = pd.DataFrame()
                df_std['ECO'] = df_c[col_eco].apply(_resolve_eco)
                df_std['Fecha'] = df_c['__fecha_std']
                df_std['Concepto'] = df_c[col_concepto] if col_concepto else df_c.get('Concepto', 'COMBUSTIBLE')
                df_std['Tipo'] = df_c[col_tipo] if col_tipo else df_c.get('Tipo', None)
                df_std['Importe'] = pd.to_numeric(df_c[col_importe], errors='coerce')
                df_std['Sistema'] = sistema
                df_std['Cantidad'] = pd.to_numeric(df_c[col_cantidad], errors='coerce') if col_cantidad else pd.Series([None] * len(df_c))
                
                df_std = df_std.dropna(subset=['Importe', 'Fecha', 'ECO'])
                # Filter ECOs using regex to only allow AU and CA (and optional LZC)
                df_std = df_std[df_std['ECO'].astype(str).str.match(r'^(AU|CA)-\d{3}(?:\s*LZC)?$', na=False)].copy()
                
                # Eliminar transacciones duplicadas provenientes de respaldos solapados
                dedup_cols = ['ECO', 'Fecha', 'Concepto', 'Tipo', 'Importe', 'Sistema', 'Cantidad']
                df_std = df_std.drop_duplicates(subset=dedup_cols).copy()
                lista_consumos.append(df_std)
                print(f"✅ Cargados {len(df_std)} registros de utilitarios desde {archivo} ({sistema}).")
            except Exception as e:
                print(f"⚠️ Error al leer {archivo}: {e}")

    # 3. Formatear y alinear Mantenimientos
    df_mantenimientos = pd.DataFrame()
    if not df_mantenimientos_raw.empty:
        try:
            # Buscar columnas necesarias
            col_eco = next((c for c in df_mantenimientos_raw.columns if 'eco' in c.lower()), None)
            col_fecha = next((c for c in df_mantenimientos_raw.columns if 'fecha' in c.lower()), None)
            col_importe = next((c for c in df_mantenimientos_raw.columns if any(kw in c.lower() for kw in ['importe', 'precioneto', 'costo'])), None)
            
            if col_eco and col_fecha and col_importe:
                df_mantenimientos['ECO'] = df_mantenimientos_raw[col_eco].apply(_normalize_eco)
                df_mantenimientos['Fecha'] = pd.to_datetime(df_mantenimientos_raw[col_fecha], errors='coerce')
                df_mantenimientos['Concepto'] = 'MANTENIMIENTO'
                df_mantenimientos['Tipo'] = None
                df_mantenimientos['Importe'] = pd.to_numeric(df_mantenimientos_raw[col_importe], errors='coerce')
                df_mantenimientos['Sistema'] = 'Excel Mantenimientos'
                df_mantenimientos['Cantidad'] = None
                df_mantenimientos = df_mantenimientos.dropna(subset=['Importe', 'Fecha', 'ECO'])
                df_mantenimientos = df_mantenimientos[df_mantenimientos['ECO'].str.match(r'^(AU|CA)-\d{3}(?:\s*LZC)?$', na=False)].copy()
                df_mantenimientos = df_mantenimientos.drop_duplicates().copy()
                print(f"✅ Procesados {len(df_mantenimientos)} registros de Mantenimiento local.")
            else:
                print("⚠️ No se pudieron identificar las columnas requeridas (ECO, Fecha, Importe/PrecioNeto) en Mantenimientos.")
        except Exception as e:
            print(f"⚠️ Error al alinear la estructura de Mantenimientos: {e}")

    # 4. UNION ALL (Concatenación)
    todas_fuentes = lista_consumos + ([df_mantenimientos] if not df_mantenimientos.empty else [])
    if not todas_fuentes:
        print("❌ Error: No se encontraron datos de consumos ni de mantenimientos para unificar.")
        return
        
    df_consumos_total = pd.concat(todas_fuentes, ignore_index=True)
    
    # 5. Cruce con Tabla Maestra (FULL OUTER JOIN por ECO normalizado)
    df_maestra['ECO'] = df_maestra['ECO'].apply(_normalize_eco)
    df_maestra['ECO_key'] = df_maestra['ECO'].apply(_clean_eco_key)
    df_consumos_total['ECO_key'] = df_consumos_total['ECO'].apply(_clean_eco_key)
    
    # Hacer el merge
    df_merge = pd.merge(df_maestra, df_consumos_total, on='ECO_key', how='outer', suffixes=('_maestra', '_consumo'))
    
    # Resolver columna ECO unificada (COALESCE en SQL)
    df_merge['ECO'] = df_merge['ECO_maestra'].fillna(df_merge['ECO_consumo'])
    df_merge = df_merge.drop(columns=['ECO_maestra', 'ECO_consumo', 'ECO_key'])
    
    # Rellenar valores nulos para columnas de importe y concepto
    df_merge['Importe'] = df_merge['Importe'].fillna(0)
    df_merge['Concepto'] = df_merge['Concepto'].fillna('SIN ACTIVIDAD')
    df_merge['Sistema'] = df_merge['Sistema'].fillna('N/A')
    
    # 6. Crear Columnas de Dashboard e Indicadores
    today = pd.Timestamp.now()
    df_merge['Fecha'] = pd.to_datetime(df_merge['Fecha'], errors='coerce')
    df_merge['Anio'] = df_merge['Fecha'].dt.year
    df_merge['Mes'] = df_merge['Fecha'].dt.month
    df_merge['Semana_Mes'] = ((df_merge['Fecha'].dt.day - 1) // 7 + 1).fillna(0).astype(int)
    
    # es_ytd: Año actual y mes menor/igual al actual, o años anteriores
    df_merge['es_ytd'] = ((df_merge['Anio'] == today.year) & (df_merge['Mes'] <= today.month)) | (df_merge['Anio'] < today.year)
    
    # incluir_en_kpi
    df_merge['incluir_en_kpi'] = df_merge['Mes'] <= today.month
    
    # Nombre_Mes
    meses_nombres = {
        1: '01 - Enero', 2: '02 - Febrero', 3: '03 - Marzo', 4: '04 - Abril',
        5: '05 - Mayo', 6: '06 - Junio', 7: '07 - Julio', 8: '08 - Agosto',
        9: '09 - Septiembre', 10: '10 - Octubre', 11: '11 - Noviembre', 12: '12 - Diciembre'
    }
    df_merge['Nombre_Mes'] = df_merge['Mes'].map(meses_nombres)
    
    # Ajustes finales de columnas
    df_merge = df_merge.drop(columns=['Centro de Trabajo', 'Domicilio'], errors='ignore')
    df_merge = df_merge.rename(columns={'Cantidad': 'Litros'})
    
    # Eliminar 'Litros' = 0 para que no salga en Pase (peajes)
    if 'Litros' in df_merge.columns:
        df_merge['Litros'] = df_merge['Litros'].replace(0, pd.NA)
        
    # Mover ECO al inicio
    cols = df_merge.columns.tolist()
    if 'ECO' in cols:
        cols.insert(0, cols.pop(cols.index('ECO')))
        df_merge = df_merge[cols]
    
    # 7. Exportar o anexar a Excel Final
    output_path = os.getenv('EXCEL_OUTPUT_PATH')
    if not output_path:
        output_dir = os.path.join(PROJECT_ROOT, "Reportes_Ejecutable")
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, "Reporte_Dashboard_Final.xlsx")
    else:
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

    def _get_signatures(df):
        ecos = df['ECO'].fillna('').astype(str).str.strip()
        fechas = pd.to_datetime(df['Fecha'], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
        conceptos = df['Concepto'].fillna('').astype(str).str.strip().str.upper()
        importes = pd.to_numeric(df['Importe'], errors='coerce').fillna(0).round(2).astype(str)
        sistemas = df['Sistema'].fillna('').astype(str).str.strip().str.upper()
        litros = pd.to_numeric(df['Litros'], errors='coerce').fillna(0).round(2).astype(str)
        return ecos + "_" + fechas + "_" + conceptos + "_" + importes + "_" + sistemas + "_" + litros

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        # Estilos comunes
        century_font = Font(name="Century Gothic")
        century_bold = Font(name="Century Gothic", bold=True)
        middle_align = Alignment(vertical="center")
        green_fill = PatternFill(start_color="E4EDEC", end_color="E4EDEC", fill_type="solid")

        excel_exists = os.path.exists(output_path)
        datos_existe = False
        
        if excel_exists:
            try:
                # keep_vba=True is only required/valid for .xlsm files; using it on .xlsx corrupts the file structure
                is_xlsm = output_path.lower().endswith('.xlsm')
                wb = openpyxl.load_workbook(output_path, keep_vba=is_xlsm)
                if wb is not None and 'Datos' in wb.sheetnames:
                    datos_existe = True
                else:
                    if wb is not None:
                        wb.close()
            except Exception as e:
                print(f"⚠️ Error al abrir el Excel existente ({e}). Se creará de nuevo.")
                excel_exists = False

        if excel_exists and datos_existe:
            print("💾 Excel existente encontrado. Realizando ingesta incremental...")
            try:
                df_existing = pd.read_excel(output_path, sheet_name='Datos')
                has_semana_mes = 'Semana_Mes' in df_existing.columns
                
                if not has_semana_mes:
                    print("🔄 Detectada estructura anterior de Datos (sin Semana_Mes). Regenerando la pestaña completa con las nuevas columnas helper...")
                    datos_existe = False
                    df_new = df_merge
                else:
                    existing_sigs = set(_get_signatures(df_existing))
                    merge_sigs = _get_signatures(df_merge)
                    df_new = df_merge[~merge_sigs.isin(existing_sigs)].copy()
            except Exception as e:
                print(f"⚠️ Error al leer datos existentes ({e}). Se reescribirá el archivo completo.")
                excel_exists = False
                df_new = df_merge

        if not excel_exists or not datos_existe:
            df_new = df_merge
            if not excel_exists:
                print("🆕 Creando nuevo archivo Excel consolidado...")
                wb = openpyxl.Workbook()
                default_sheet = wb.active
                wb.remove(default_sheet)
                ws_datos = wb.create_sheet(title='Datos')
            else:
                print("🔄 Regenerando pestaña Datos...")
                if 'Datos' in wb.sheetnames:
                    del wb['Datos']
                ws_datos = wb.create_sheet(title='Datos')
            
            # Escribir encabezados
            headers = df_new.columns.tolist()
            ws_datos.append(headers)
            ws_datos.row_dimensions[1].height = 20
            for col_idx, header in enumerate(headers, 1):
                cell = ws_datos.cell(row=1, column=col_idx)
                cell.font = century_bold
                cell.fill = green_fill
                cell.alignment = middle_align
        else:
            ws_datos = wb['Datos']

        # Escribir filas nuevas
        if not df_new.empty:
            print(f"📥 Insertando {len(df_new)} nuevos registros...")
            start_row = ws_datos.max_row + 1
            
            # openpyxl append row
            for r_idx, row in df_new.iterrows():
                row_vals = []
                for col_name in df_merge.columns:
                    val = row[col_name]
                    # Convert pandas NA/NaT/NaN to None for openpyxl
                    if pd.isna(val):
                        val = None
                    row_vals.append(val)
                ws_datos.append(row_vals)
            
            # Formatear solo las filas nuevas agregadas (Optimizado con iter_rows)
            end_row = ws_datos.max_row
            if (end_row - start_row) <= 5000:
                print(f"🎨 Aplicando formato a {end_row - start_row + 1} filas nuevas...")
                for row in ws_datos.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=ws_datos.max_column):
                    ws_datos.row_dimensions[row[0].row].height = 20
                    for cell in row:
                        cell.font = century_font
                        cell.alignment = middle_align
            else:
                print("⚡ Muchos registros nuevos. Omitiendo formato de celdas individuales para agilizar el proceso.")
        else:
            print("✨ No se encontraron registros nuevos para añadir.")

        ws_datos.sheet_view.showGridLines = False

        # Autoajustar anchos de columnas en Datos
        for col in ws_datos.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            ws_datos.column_dimensions[col_letter].width = max(max_len + 3, 10)

        # 8. Ingesta de Movimientos Fuera de Horario Laboral
        mov_path = os.getenv('EXCEL_MOV_NOLABORALES_PATH')
        if mov_path and os.path.exists(mov_path):
            print("🚗 Procesando movimientos fuera de horario laboral...")
            try:
                # Leer pestaña Historico
                df_mov_raw = pd.read_excel(mov_path, sheet_name='Historico')
                # Normalizar ECO
                df_mov_raw['ECO'] = df_mov_raw['ECO'].apply(_normalize_eco)
                # Filtrar ECOs válidos
                df_mov_raw = df_mov_raw[df_mov_raw['ECO'].str.match(r'^(AU|CA)-\d{3}(?:\s*LZC)?$', na=False)].copy()
                
                # Definir firmas de movimientos
                def _get_mov_sigs(df):
                    ecos = df['ECO'].fillna('').astype(str).str.strip()
                    f_ini = pd.to_datetime(df['Fecha-hora Inicio'], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
                    f_fin = pd.to_datetime(df['Fecha-hora Término'], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S').fillna('')
                    usuarios = df['Usuario'].fillna('').astype(str).str.strip().str.upper()
                    distancias = pd.to_numeric(df['Distancia(KM)'], errors='coerce').fillna(0).round(2).astype(str)
                    return ecos + "_" + f_ini + "_" + f_fin + "_" + usuarios + "_" + distancias

                def _compute_helper_columns(df):
                    if df.empty:
                        return df.copy()
                    df_out = df.copy()
                    dt_series = pd.to_datetime(df_out['Fecha-hora Inicio'], errors='coerce')
                    df_out['Anio'] = dt_series.dt.year.fillna(0).astype(int)
                    df_out['Mes_Num'] = dt_series.dt.month.fillna(0).astype(int)
                    df_out['Semana_Mes'] = ((dt_series.dt.day - 1) // 7 + 1).fillna(0).astype(int)
                    
                    df_out['Fecha_Solo'] = dt_series.dt.date
                    dow_map = {
                        0: 'Lunes', 1: 'Martes', 2: 'Miércoles', 3: 'Jueves', 4: 'Viernes',
                        5: 'Sábado', 6: 'Domingo'
                    }
                    dia_nombre = dt_series.dt.dayofweek.map(dow_map).fillna('')
                    hora_frac = dt_series.dt.hour + dt_series.dt.minute / 60.0
                    
                    is_dom = dia_nombre == 'Domingo'
                    is_sab_tarde = (dia_nombre == 'Sábado') & (hora_frac >= 16.0)
                    is_sem_noche = (~dia_nombre.isin(['Sábado', 'Domingo'])) & (hora_frac >= 20.0)
                    
                    is_dom_first = is_dom & ~df_out[is_dom].duplicated(subset=['ECO', 'Fecha_Solo'])
                    df_out['Es_Domingo_Aux'] = is_dom_first.reindex(df_out.index, fill_value=False).astype(int)
                    
                    is_sab_tarde_first = is_sab_tarde & ~df_out[is_sab_tarde].duplicated(subset=['ECO', 'Fecha_Solo'])
                    df_out['Es_Sabado_Tarde_Aux'] = is_sab_tarde_first.reindex(df_out.index, fill_value=False).astype(int)
                    
                    is_sem_noche_first = is_sem_noche & ~df_out[is_sem_noche].duplicated(subset=['ECO', 'Fecha_Solo'])
                    df_out['Es_Semana_Noche_Aux'] = is_sem_noche_first.reindex(df_out.index, fill_value=False).astype(int)
                    
                    df_out.drop(columns=['Fecha_Solo'], inplace=True)
                    return df_out

                mov_sheet_exists = 'Movimientos' in wb.sheetnames
                df_new_mov = pd.DataFrame()
                has_helpers = False

                if mov_sheet_exists:
                    try:
                        df_existing_mov = pd.read_excel(output_path, sheet_name='Movimientos')
                        has_helpers = 'Anio' in df_existing_mov.columns
                        existing_mov_sigs = set(_get_mov_sigs(df_existing_mov))
                        new_mov_sigs = _get_mov_sigs(df_mov_raw)
                        df_new_mov = df_mov_raw[~new_mov_sigs.isin(existing_mov_sigs)].copy()
                    except Exception as e:
                        print(f"⚠️ Error al leer movimientos existentes ({e}). Se reescribirá la pestaña.")
                        mov_sheet_exists = False
                        df_new_mov = df_mov_raw
                else:
                    df_new_mov = df_mov_raw

                # Caso 1: La pestaña ya existe pero NO tiene las columnas helper (Backfill)
                if mov_sheet_exists and not has_helpers:
                    print("🔄 Detectada estructura anterior de Movimientos. Regenerando la pestaña completa con las nuevas columnas helper...")
                    df_all = pd.concat([df_existing_mov, df_new_mov], ignore_index=True)
                    df_all = _compute_helper_columns(df_all)
                    
                    ws_mov = wb['Movimientos']
                    ws_mov.delete_rows(1, ws_mov.max_row + 1)
                    headers_mov = df_all.columns.tolist()
                    ws_mov.append(headers_mov)
                    ws_mov.row_dimensions[1].height = 20
                    for col_idx, header in enumerate(headers_mov, 1):
                        cell = ws_mov.cell(row=1, column=col_idx)
                        cell.font = century_bold
                        cell.fill = green_fill
                        cell.alignment = middle_align
                        
                    start_row_mov = 2
                    for r_idx, row in df_all.iterrows():
                        row_vals = []
                        for col_name in df_all.columns:
                            val = row[col_name]
                            if pd.isna(val):
                                val = None
                            row_vals.append(val)
                        ws_mov.append(row_vals)
                    end_row_mov = ws_mov.max_row
                    
                    if (end_row_mov - start_row_mov) <= 5000:
                        print(f"🎨 Aplicando formato a {end_row_mov - start_row_mov + 1} filas de movimientos...")
                        for row in ws_mov.iter_rows(min_row=start_row_mov, max_row=end_row_mov, min_col=1, max_col=ws_mov.max_column):
                            ws_mov.row_dimensions[row[0].row].height = 20
                            for cell in row:
                                cell.font = century_font
                                cell.alignment = middle_align
                    else:
                        print("⚡ Muchos registros. Omitiendo formato de celdas individuales en Movimientos para agilizar el proceso.")

                # Caso 2: La pestaña no existe (creación inicial con helpers)
                elif not mov_sheet_exists:
                    ws_mov = wb.create_sheet(title='Movimientos')
                    df_new_mov = _compute_helper_columns(df_new_mov)
                    headers_mov = df_new_mov.columns.tolist()
                    ws_mov.append(headers_mov)
                    ws_mov.row_dimensions[1].height = 20
                    for col_idx, header in enumerate(headers_mov, 1):
                        cell = ws_mov.cell(row=1, column=col_idx)
                        cell.font = century_bold
                        cell.fill = green_fill
                        cell.alignment = middle_align
                        
                    start_row_mov = 2
                    if not df_new_mov.empty:
                        for r_idx, row in df_new_mov.iterrows():
                            row_vals = []
                            for col_name in df_new_mov.columns:
                                val = row[col_name]
                                if pd.isna(val):
                                    val = None
                                row_vals.append(val)
                            ws_mov.append(row_vals)
                        end_row_mov = ws_mov.max_row
                        
                        if (end_row_mov - start_row_mov) <= 5000:
                            print(f"🎨 Aplicando formato a {end_row_mov - start_row_mov + 1} filas de movimientos...")
                            for row in ws_mov.iter_rows(min_row=start_row_mov, max_row=end_row_mov, min_col=1, max_col=ws_mov.max_column):
                                ws_mov.row_dimensions[row[0].row].height = 20
                                for cell in row:
                                    cell.font = century_font
                                    cell.alignment = middle_align
                        else:
                            print("⚡ Muchos registros. Omitiendo formato de celdas individuales en Movimientos para agilizar el proceso.")

                # Caso 3: La pestaña ya existe y ya cuenta con las columnas helper (Append)
                else:
                    ws_mov = wb['Movimientos']
                    if not df_new_mov.empty:
                        print(f"📥 Insertando {len(df_new_mov)} nuevos registros de movimientos...")
                        df_new_mov = _compute_helper_columns(df_new_mov)
                        start_row_mov = ws_mov.max_row + 1
                        
                        for r_idx, row in df_new_mov.iterrows():
                            row_vals = []
                            for col_name in df_new_mov.columns:
                                val = row[col_name]
                                if pd.isna(val):
                                    val = None
                                row_vals.append(val)
                            ws_mov.append(row_vals)
                            
                        end_row_mov = ws_mov.max_row
                        if (end_row_mov - start_row_mov) <= 5000:
                            print(f"🎨 Aplicando formato a {end_row_mov - start_row_mov + 1} filas nuevas de movimientos...")
                            for row in ws_mov.iter_rows(min_row=start_row_mov, max_row=end_row_mov, min_col=1, max_col=ws_mov.max_column):
                                ws_mov.row_dimensions[row[0].row].height = 20
                                for cell in row:
                                    cell.font = century_font
                                    cell.alignment = middle_align
                        else:
                            print("⚡ Muchos registros. Omitiendo formato de celdas individuales en Movimientos para agilizar el proceso.")
                    else:
                        print("✨ No se encontraron movimientos nuevos para añadir.")

                ws_mov.sheet_view.showGridLines = False

                # Autoajustar anchos de columnas en Movimientos
                for col in ws_mov.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    ws_mov.column_dimensions[col_letter].width = max(max_len + 3, 10)

            except Exception as e:
                print(f"⚠️ Error al procesar movimientos fuera de horario laboral: {e}")
        else:
            print("⚠️ No se encontró la ruta del reporte de movimientos en .env o el archivo no existe.")

        # Redimensionar tablas de Excel automáticamente al número real de filas
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for table_name, table in list(ws.tables.items()):
                ref_parts = table.ref.split(':')
                if len(ref_parts) == 2:
                    start_cell = ref_parts[0]
                    end_cell = ref_parts[1]
                    m = re.match(r'^([A-Z]+)', end_cell)
                    if m:
                        end_col = m.group(1)
                        new_ref = f"{start_cell}:{end_col}{ws.max_row}"
                        table.ref = new_ref
                        print(f"📊 Tabla '{table_name}' redimensionada automáticamente a {new_ref} en la hoja '{sheet_name}'.")

        wb.save(output_path)
        wb.close()
        
        print(f"\n✅ ¡Proceso global completado exitosamente!")
        print(f"📊 Reporte Dashboard generado en: {output_path}")
    except PermissionError:
        print(f"❌ Error de permisos: El archivo de salida '{output_path}' está siendo usado por otro programa (probablemente está abierto en Excel). Por favor, ciérralo e intenta de nuevo.")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"❌ Error crítico al exportar el archivo de reporte final: {e}")

if __name__ == "__main__":
    consolidar_todo()
